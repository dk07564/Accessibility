import matplotlib.pyplot as plt
from matplotlib.pyplot import figure
#from AccessibilityTest import *
from AccessibilityTest import *
import numpy
import openpyxl

xPlt=numpy.arange(len(passList))
plt.bar(xPlt, widgetList, label="widget")
plt.plot(xPlt, passList, label="passed widget", color="r", marker=".")

plt.xlabel("xml")
plt.ylabel("widget")
plt.xticks(xPlt, xmlList, fontsize=7)
plt.legend()
plt.savefig(name+".png")

wb=openpyxl.Workbook()
sheet=wb.active

excelAttribute=["파일 이름", "평가한 위젯 수", "통과된 위젯 수"]

for i in range(1, len(excelAttribute)+1):
    sheet.cell(row=1, column=i).value=excelAttribute[i-1]

for i in range(2, len(xmlList)+2):
    for j in range(1, 4):
        if(j==1):
            sheet.cell(row=i, column=j).value = accList[i-2]
        elif(j==2):
            sheet.cell(row=i, column=j).value = widgetList[i-2]
        else:
            sheet.cell(row=i, column=j).value = passList[i-2]

wb.save(name+".xlsx")